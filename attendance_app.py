# attendance_app.py (updated with oval blue buttons + black-bordered white input boxes)

import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook, load_workbook
import os
import glob
from attendance_utils import get_students, mark_attendance, merge_all_students, get_absentees

# GUI Setup
root = tk.Tk()
root.title("Attendance System")
root.geometry("500x800")
root.configure(bg="#f0f8ff")

# Define style
style = ttk.Style()
style.theme_use("clam")
style.configure("Rounded.TButton",
                font=("Arial", 11, "bold"),
                padding=10,
                borderwidth=1,
                relief="flat",
                background="#4682B4",
                foreground="white",
                anchor="center",
                bordercolor="#4682B4")
style.map("Rounded.TButton",
          background=[("active", "#4169E1")],
          foreground=[("disabled", "gray")])

# Oval button wrapper
def create_oval_button(text, command):
    return tk.Button(root, text=text, command=command, font=("Arial", 11, "bold"), bg="#4682B4", fg="white",
                     relief="flat", bd=0, highlightthickness=0, padx=20, pady=10, cursor="hand2")

# Bordered entry
def create_bordered_entry():
    frame = tk.Frame(root, bg="black", padx=1, pady=1)
    entry = tk.Entry(frame, font=("Arial", 12), bg="white", fg="black", relief="flat", width=30)
    entry.pack()
    frame.pack()
    return entry

# Custom popup
def show_custom_popup(title, message, color="#007acc"):
    popup = tk.Toplevel()
    popup.title(title)
    popup.geometry("300x150")
    popup.configure(bg=color)
    tk.Label(popup, text=message, bg=color, fg="white", font=("Arial", 12, "bold")).pack(pady=30)
    ttk.Button(popup, text="OK", command=popup.destroy, style="Rounded.TButton").pack()

# Functions

def get_class_files():
    return [f for f in glob.glob("*.xlsx") if f != "all_students.xlsx"]

def refresh_dropdown():
    class_dropdown['values'] = get_class_files()
    if get_class_files():
        class_var.set(get_class_files()[0])

def add_student_to_excel(name, file_name):
    if not name:
        show_custom_popup("Warning", "Please enter a student name.", color="#87ceeb")
        return
    if not file_name:
        show_custom_popup("Warning", "Please select a class file.", color="#87ceeb")
        return
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.cell(row=1, column=1).value = "Name"
        wb.save(file_name)
    wb = load_workbook(file_name)
    ws = wb.active
    row = 2
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    ws.cell(row=row, column=1).value = name
    wb.save(file_name)
    wb.close()
    show_custom_popup("Success", f"{name} added to {file_name}", color="#4682B4")

def create_class_file():
    name = new_class_entry.get().strip()
    if not name.endswith(".xlsx"):
        name += ".xlsx"
    if os.path.exists(name):
        show_custom_popup("Exists", f"{name} already exists.", color="#87ceeb")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.cell(row=1, column=1).value = "Name"
        wb.save(name)
        show_custom_popup("Created", f"{name} created successfully.", color="#4682B4")
    new_class_entry.delete(0, tk.END)
    refresh_dropdown()

def delete_class_file():
    file = class_var.get()
    if file and os.path.exists(file):
        try:
            os.remove(file)
            show_custom_popup("Deleted", f"{file} has been deleted.", color="#4682B4")
            refresh_dropdown()
        except Exception as e:
            show_custom_popup("Error", str(e), color="#87ceeb")
    else:
        show_custom_popup("Error", "Select a valid file to delete.", color="#87ceeb")

def add_name():
    name = add_entry.get().strip()
    file = class_var.get().strip()
    add_student_to_excel(name, file)
    add_entry.delete(0, tk.END)

checkbox_frame = tk.Frame(root, bg="#f0f8ff")
checkboxes = {}

def load_students():
    for widget in checkbox_frame.winfo_children():
        widget.destroy()
    checkboxes.clear()
    selected_file = class_var.get()
    if not selected_file:
        show_custom_popup("Error", "Please select a file.", color="#87ceeb")
        return
    students = get_students(selected_file)
    if not students:
        show_custom_popup("No Data", "No students found in selected file.", color="#87ceeb")
        return
    tk.Label(checkbox_frame, text="Select Present Students:", font=("Arial", 12, "bold"), bg="#f0f8ff", fg="#333").pack()
    for student in students:
        var = tk.IntVar()
        cb = tk.Checkbutton(checkbox_frame, text=student, variable=var, font=("Arial", 10), bg="#f0f8ff")
        cb.pack(anchor="w")
        checkboxes[student] = var

def submit_attendance():
    selected_file = class_var.get()
    if not selected_file:
        show_custom_popup("Error", "Please select a file.", color="#87ceeb")
        return
    present = [name for name, var in checkboxes.items() if var.get() == 1]
    result = mark_attendance(selected_file, present)
    total = len(checkboxes)
    present_count = len(present)
    absent_count = total - present_count
    if result == "Marked successfully":
        show_custom_popup("Attendance Done", f"Present: {present_count}\nAbsent: {absent_count}", color="#4682B4")
    elif result == "Attendance already marked":
        show_custom_popup("Already Marked", "You already marked today.", color="#87ceeb")
    else:
        show_custom_popup("Error", result, color="#87ceeb")
    for var in checkboxes.values():
        var.set(0)

def merge_students():
    output_file = merge_all_students()
    show_custom_popup("Merged", f"All students saved to {output_file}", color="#4682B4")

def check_absentees():
    selected_file = class_var.get()
    date = absent_entry.get().strip()
    if not selected_file or not date:
        show_custom_popup("Missing Info", "Select a file and enter date.", color="#87ceeb")
        return
    absentees = get_absentees(selected_file, date)
    if not absentees:
        show_custom_popup("All Present", f"Everyone was present on {date}.", color="#4682B4")
    elif absentees == ["Date not found in sheet."]:
        show_custom_popup("Error", f"{date} not found in {selected_file}.", color="#87ceeb")
    else:
        show_custom_popup("Absentees", f"Absent on {date}:" + "\n".join(absentees), color="#87ceeb")

# UI Layout

# Labels & Inputs
tk.Label(root, text="Create New Class File:", font=("Arial", 12, "bold"), bg="#f0f8ff", fg="#333").pack(pady=10)
new_class_entry = create_bordered_entry()
create_oval_button("Create Class", create_class_file).pack(pady=5)
create_oval_button("Delete", delete_class_file).pack(pady=5)

tk.Label(root, text="Select Class File:", font=("Arial", 12, "bold"), bg="#f0f8ff", fg="#333").pack(pady=10)
class_var = tk.StringVar()
class_dropdown = ttk.Combobox(root, textvariable=class_var, values=get_class_files(), state="readonly", width=30)
class_dropdown.pack()
refresh_dropdown()

tk.Label(root, text="Enter New Student Name:", font=("Arial", 12, "bold"), bg="#f0f8ff", fg="#333").pack(pady=10)
add_entry = create_bordered_entry()

tk.Label(root, text="Check Absentees for Date (YYYY-MM-DD):", font=("Arial", 12, "bold"), bg="#f0f8ff", fg="#333").pack(pady=10)
absent_entry = create_bordered_entry()

# Buttons
create_oval_button("Add Student", add_name).pack(pady=5)
create_oval_button("Load Students", load_students).pack(pady=5)
create_oval_button("Mark Attendance", submit_attendance).pack(pady=5)
create_oval_button("Merge All Students", merge_students).pack(pady=10)
create_oval_button("Check Absentees", check_absentees).pack(pady=10)

checkbox_frame.pack(pady=10)

root.mainloop()
